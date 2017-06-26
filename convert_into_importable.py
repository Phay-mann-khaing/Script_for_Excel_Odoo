import openpyxl


class Convert2importable:
    def __init__(self, filename=None):
        self.filename = filename
        self.convert_function()

    def convert_function(self):
        wb = openpyxl.load_workbook(self.filename)
        wb1 = openpyxl.Workbook()  # new file
        ws = wb.active
        ws1 = wb1.active  # open sheet
        column_count = ws.max_column  # get column count
        row_count = 1  # initialize, use this as row number to write
        max_current_row = 2  # initialize, use to check max row number
        a = list(ws.iter_rows())  # loop rows
        for index1, line in enumerate(a, 1):  # loop over rows tuple
            for index2, piece in enumerate(line, 1):  # loop inside one row
                if index1 == 1:  # for header
                    if index2 < 4:  # don't write location headers
                        ws1.cell(row=index1, column=index2).value = piece.value
                    else:
                        break
                else:  # for body of rows
                    if index2 > 3:
                        break
                    count = column_count - 3
                    if count + 1 == index2:  # don't write location
                        break
                    check_null = 0
                    row = 0
                    row = row_count

                    while count > 0:  # for same product that has quantities in different location
                        if line[3 + check_null].value:  # check location to write new row for it
                            ws1.cell(row=row, column=index2 + 2).value = piece.value
                            ws1.cell(row=row, column=6).value = line[3 + check_null].value
                            a = line[3 + check_null].column
                            ws1.cell(row=row, column=7).value = ws['%s1' % a].value
                            row += 1
                        count -= 1
                        check_null += 1

                    if max_current_row < row:
                        max_current_row = row

            row_count = max_current_row

        # col_a = ws['A']  # 0-indexing
        # for idx, cell in enumerate(col_a, 1):
        #     ws1.cell(row=idx, column=1).value = cell.value

        # for cell in col_a:
        # ws.cell(row=idx, column=4).value = cell.value # 1-indexing


        # ws1.append(
        #     ['line_ids/product_qty', 'line_ids/location_id/id', 'line_ids/product_id/id', 'line_ids/product_uom/id'])

        # Change Headers
        ws1['A1'] = 'name'
        ws1['B1'] = 'location_id/id'
        ws1['C1'] = 'line_ids/product_id/id'
        ws1['D1'] = 'Product Name'
        ws1['E1'] = 'line_ids/product_uom/id'
        ws1['F1'] = 'line_ids/product_qty'
        ws1['G1'] = 'line_ids/location_id/id'
        # Save file
        wb1.save("ToImport.xlsx")
        print "Flie created as 'TestCreate.xlsx'"

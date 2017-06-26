import csv
import openpyxl
import xlrd
from openpyxl.workbook import Workbook
import convert_into_importable


class Convert2xlsx:

    def __init__(self, filename=None, is_csv=None, is_xls=None):
        self.filename = filename
        self.csv = is_csv
        self.xls = is_xls
        if self.csv == 1:
            self.cvt_csv_to_xlsx()
        elif self.xls == 1:
            self.cvt_xls_to_xlsx()

    def cvt_csv_to_xlsx(self):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active

            f = open(self.filename)
            reader = csv.reader(f, delimiter=',')
            for row in reader:
                ws.append(row)

            f.close()
            new_xlsx = 'Convertedtoxlsx.xlsx'
            wb.save(new_xlsx)
            self.filename = new_xlsx
            convert_into_importable.Convert2importable(new_xlsx)
            # return new_xlsx

        except ValueError:
            print 'File invalid'

    def cvt_xls_to_xlsx(self):
        book_xls = xlrd.open_workbook(self.filename)
        book_xlsx = Workbook()

        sheet_names = book_xls.sheet_names()
        for sheet_index in range(0, len(sheet_names)):
            sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])
            if sheet_index == 0:
                sheet_xlsx = book_xlsx.active
                sheet_xlsx.title = sheet_names[sheet_index]
            else:
                sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

            for row in range(0, sheet_xls.nrows):
                for col in range(0, sheet_xls.ncols):
                    sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

        new_xlsx = 'Convertedtoxlsx.xlsx'
        book_xlsx.save(new_xlsx)
        convert_into_importable.Convert2importable(new_xlsx)
        # return new_xlsx
